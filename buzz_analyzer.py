"""TwitterAPI.ioを使ってAI系バズポストを取得しExcelに保存するスクリプト"""

import os
import sys
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# .envファイルから環境変数を読み込む
load_dotenv()


def fetch_tweets_by_keyword(keyword, api_key, url, headers):
    """指定されたキーワードでツイートを取得"""
    query = f"{keyword} lang:ja min_faves:100"
    params = {
        "query": query,
        "queryType": "Latest",
    }

    print(f"  キーワード「{keyword}」で検索中...")

    try:
        response = requests.get(url, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
    except requests.exceptions.ConnectionError:
        print("エラー: APIサーバーに接続できません。ネットワーク接続を確認してください。")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print("エラー: APIリクエストがタイムアウトしました。")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        if response.status_code == 401:
            print("エラー: APIキーが無効です。正しいキーを設定してください。")
        elif response.status_code == 429:
            print("エラー: APIのレート制限に達しました。しばらく待ってから再実行してください。")
        else:
            print(f"エラー: APIリクエストに失敗しました (HTTP {response.status_code}): {e}")
        sys.exit(1)
    except Exception as e:
        print(f"エラー: 予期しないエラーが発生しました: {e}")
        sys.exit(1)

    tweets = data.get("tweets", [])
    print(f"    → {len(tweets)}件取得")
    return tweets


def fetch_buzz_posts():
    api_key = os.environ.get("TWITTER_API_KEY")
    if not api_key:
        print("エラー: 環境変数 TWITTER_API_KEY が設定されていません。")
        print("設定例: export TWITTER_API_KEY='your-api-key'")
        sys.exit(1)

    url = "https://api.twitterapi.io/twitter/tweet/advanced_search"
    headers = {"X-API-Key": api_key}

    # 検索キーワードリスト
    keywords = [
        "AI 副業",
        "AI 稼ぐ",
        "AI 収益化",
        "Claude Code",
        "AI自動化",
    ]

    print("バズポストを取得中...")
    print(f"検索キーワード数: {len(keywords)}個")

    # 全キーワードで検索して結果を集約
    all_tweets = []
    seen_tweet_ids = set()  # 重複排除用

    for i, keyword in enumerate(keywords):
        tweets = fetch_tweets_by_keyword(keyword, api_key, url, headers)

        # 重複排除しながら追加
        for tweet in tweets:
            tweet_id = tweet.get("id", "")
            if tweet_id and tweet_id not in seen_tweet_ids:
                seen_tweet_ids.add(tweet_id)
                all_tweets.append(tweet)

        # レート制限対策：最後のキーワード以外は待機
        if i < len(keywords) - 1:
            print("    レート制限対策で3秒待機中...")
            time.sleep(3)

    if not all_tweets:
        print("該当するポストが見つかりませんでした。")
        sys.exit(0)

    print(f"\n重複排除後の合計取得件数: {len(all_tweets)}件")

    # データ整形
    rows = []
    for tweet in all_tweets:
        user = tweet.get("author", {})
        tweet_id = tweet.get("id", "")
        username = user.get("userName", "")
        post_url = f"https://x.com/{username}/status/{tweet_id}" if username and tweet_id else ""

        rows.append({
            "本文": tweet.get("text", ""),
            "いいね数": tweet.get("likeCount", 0),
            "リポスト数": tweet.get("retweetCount", 0),
            "リプライ数": tweet.get("replyCount", 0),
            "投稿日時": tweet.get("createdAt", ""),
            "ユーザー名": username,
            "フォロワー数": "未取得",
            "ポストURL": post_url,
        })

    # いいね数で降順ソート
    rows.sort(key=lambda x: x["いいね数"], reverse=True)

    # outputフォルダを作成
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    # Excelファイル作成
    today = datetime.now().strftime("%Y%m%d")
    filename = os.path.join(output_dir, f"buzz_posts_{today}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "バズポスト"

    # ヘッダー
    headers = ["本文", "いいね数", "リポスト数", "リプライ数", "投稿日時", "ユーザー名", "フォロワー数", "ポストURL"]
    ws.append(headers)

    # ヘッダー行のスタイル設定
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行を追加
    for row_data in rows:
        ws.append([
            row_data["本文"],
            row_data["いいね数"],
            row_data["リポスト数"],
            row_data["リプライ数"],
            row_data["投稿日時"],
            row_data["ユーザー名"],
            row_data["フォロワー数"],
            row_data["ポストURL"],
        ])

    # 本文列（A列）の折り返し設定
    for row in range(2, len(rows) + 2):
        cell = ws.cell(row=row, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 列幅の自動調整
    column_widths = {
        1: 60,  # 本文
        2: 12,  # いいね数
        3: 12,  # リポスト数
        4: 12,  # リプライ数
        5: 25,  # 投稿日時
        6: 20,  # ユーザー名
        7: 12,  # フォロワー数
        8: 50,  # ポストURL
    }

    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # オートフィルター設定
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # 保存
    wb.save(filename)
    print(f"保存完了: {filename}")


if __name__ == "__main__":
    fetch_buzz_posts()
