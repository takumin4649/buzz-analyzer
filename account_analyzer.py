"""特定のXアカウントの投稿を取得して、アカウントごとの傾向を分析するスクリプト"""

import os
import sys
import time
import re
import argparse
from datetime import datetime, timedelta
from collections import Counter
from statistics import mean, median

# Windows環境での文字化け対策
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# .envファイルから環境変数を読み込む
load_dotenv()

# 対象アカウントリスト
TARGET_ACCOUNTS = [
    "1banana2546",
    "Naoki_GPT",
    "ethann_AI",
    "maind200",
    "gyarados__AI",
    "ComagerTon79278",
    "mamiya_afi",
    "takkun_life_ea",
    "ai_nepro",
]


def load_test_data_from_excel(excel_path):
    """テストモード用：既存のExcelファイルからデータを読み込む"""
    print(f"📂 テストモード: {excel_path} からデータを読み込み中...")

    if not os.path.exists(excel_path):
        print(f"エラー: ファイルが見つかりません: {excel_path}")
        return {}

    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]  # 最初のシートを使用

    # ヘッダー行を読み込み
    headers = [cell.value for cell in ws[1]]

    # カラム名のマッピング（柔軟に対応）
    column_mapping = {}
    for i, header in enumerate(headers, 1):
        if header in ["本文", "text", "テキスト"]:
            column_mapping["text"] = i
        elif header in ["いいね数", "likeCount", "likes"]:
            column_mapping["likeCount"] = i
        elif header in ["リポスト数", "RT数", "retweetCount", "retweets"]:
            column_mapping["retweetCount"] = i
        elif header in ["リプライ数", "replyCount", "replies"]:
            column_mapping["replyCount"] = i
        elif header in ["投稿日時", "createdAt", "created_at"]:
            column_mapping["createdAt"] = i
        elif header in ["ユーザー名", "userName", "username", "user_name"]:
            column_mapping["userName"] = i
        elif header in ["ポストURL", "投稿URL", "url", "tweet_url"]:
            column_mapping["url"] = i

    print(f"  カラムマッピング: {column_mapping}")

    # データを読み込み（ユーザーごとにグループ化）
    tweets_by_user = {}
    total_tweets = 0

    for row in range(2, ws.max_row + 1):
        # 各カラムの値を取得
        text = ws.cell(row, column_mapping.get("text", 1)).value or ""
        like_count = ws.cell(row, column_mapping.get("likeCount", 2)).value or 0
        retweet_count = ws.cell(row, column_mapping.get("retweetCount", 3)).value or 0
        reply_count = ws.cell(row, column_mapping.get("replyCount", 4)).value or 0
        created_at = ws.cell(row, column_mapping.get("createdAt", 5)).value or ""
        username = ws.cell(row, column_mapping.get("userName", 6)).value or "unknown"
        post_url = ws.cell(row, column_mapping.get("url", 8)).value or ""

        # ユーザー名を正規化（@を除去）
        if isinstance(username, str):
            username = username.lstrip("@")

        # ポストURLからIDを抽出
        tweet_id = ""
        if post_url:
            match = re.search(r"/status/(\d+)", post_url)
            if match:
                tweet_id = match.group(1)

        # Tweet形式に変換
        tweet = {
            "text": text,
            "likeCount": int(like_count) if isinstance(like_count, (int, float)) else 0,
            "retweetCount": int(retweet_count) if isinstance(retweet_count, (int, float)) else 0,
            "replyCount": int(reply_count) if isinstance(reply_count, (int, float)) else 0,
            "createdAt": str(created_at),
            "id": tweet_id,
            "author": {
                "userName": username,
            },
        }

        # ユーザーごとに分類
        if username not in tweets_by_user:
            tweets_by_user[username] = []
        tweets_by_user[username].append(tweet)
        total_tweets += 1

    print(f"  ✅ 読み込み完了: {total_tweets}件のツイート、{len(tweets_by_user)}アカウント")
    for username, tweets in tweets_by_user.items():
        print(f"     - @{username}: {len(tweets)}件")

    return tweets_by_user


def fetch_user_tweets(username, api_key, count=100):
    """指定されたユーザーの投稿を取得"""
    url = "https://api.twitterapi.io/twitter/user/last_tweets"
    headers = {"X-API-Key": api_key}
    params = {
        "userName": username,  # 正しいパラメータ名
        "includeReplies": False,  # リプライを除外
    }

    print(f"  @{username} の投稿を取得中...")

    all_tweets = []
    cursor = None

    # ページネーションで最大count件まで取得
    while len(all_tweets) < count:
        if cursor:
            params["cursor"] = cursor

        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()

            # エラーレスポンスのチェック
            if data.get("status") == "error":
                print(f"エラー: {data.get('message', '不明なエラー')}（@{username}）")
                return all_tweets

            # ツイートを取得（data.data.tweetsからアクセス）
            data_content = data.get("data", {})
            tweets = data_content.get("tweets", [])
            if not tweets:
                break

            all_tweets.extend(tweets)

            # 次のページがあるかチェック（トップレベルから取得）
            has_next = data.get("has_next_page", False)
            if not has_next:
                break

            cursor = data.get("next_cursor")
            if not cursor:
                break

        except requests.exceptions.ConnectionError:
            print(f"エラー: APIサーバーに接続できません（@{username}）")
            return all_tweets
        except requests.exceptions.Timeout:
            print(f"エラー: APIリクエストがタイムアウトしました（@{username}）")
            return all_tweets
        except requests.exceptions.HTTPError as e:
            if response.status_code == 401:
                print(f"エラー: APIキーが無効です（@{username}）")
            elif response.status_code == 429:
                print(f"エラー: APIのレート制限に達しました（@{username}）")
            else:
                print(f"エラー: APIリクエストに失敗しました（@{username}）: HTTP {response.status_code}")
            return all_tweets
        except Exception as e:
            print(f"エラー: 予期しないエラーが発生しました（@{username}）: {e}")
            return all_tweets

    # 必要な件数に制限
    all_tweets = all_tweets[:count]
    print(f"    → {len(all_tweets)}件取得")
    return all_tweets


def classify_opening_pattern(text):
    """冒頭パターンを分類"""
    # 改行や空白を除去して最初の文を取得
    first_line = text.strip().split("\n")[0].strip()

    # 数字提示パターン（冒頭に数字がある）
    if re.match(r"^\d+[.、。:：\s]", first_line):
        return "数字提示"

    # 疑問形パターン（？で終わる、または疑問詞で始まる）
    if "？" in first_line or "?" in first_line or re.match(r"^(何|どう|いつ|どこ|誰|なぜ|どの)", first_line):
        return "疑問形"

    # 断定形パターン（です、だ、である、ですで終わる）
    if re.search(r"(です|だ|である|ます|ました|でした)[\s。]*$", first_line):
        return "断定形"

    # 共感パターン（わかる、あるある、そう、ほんと、まじ など）
    if re.search(r"(わかる|あるある|そう|ほんと|まじ|やば|すご|えぐ)", first_line, re.IGNORECASE):
        return "共感"

    return "その他"


def detect_cta(text):
    """CTA（Call To Action）の有無を検出"""
    cta_patterns = [
        r"(やって|試して|使って|見て|読んで|チェック|クリック|登録|フォロー|リプ|RT|シェア)",
        r"(ください|してね|しよう|しましょう|おすすめ)",
        r"(こちら|リンク|プロフ|固定|bio)",
    ]

    for pattern in cta_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return True
    return False


def analyze_emotion(text):
    """感情分析（ルールベース）"""
    emotions = []

    # 期待
    if re.search(r"(楽しみ|ワクワク|期待|できる|可能|チャンス|やばい|すごい|いける)", text, re.IGNORECASE):
        emotions.append("期待")

    # 驚き
    if re.search(r"(びっくり|驚き|まさか|信じられ|ファッ|えぇ|マジ|ほんと|!|！)", text, re.IGNORECASE):
        emotions.append("驚き")

    # 共感
    if re.search(r"(わかる|あるある|そう|ほんと|同じ|私も|俺も|確かに)", text, re.IGNORECASE):
        emotions.append("共感")

    # 恐怖・不安
    if re.search(r"(怖い|不安|心配|危険|やばい|ダメ|失敗)", text, re.IGNORECASE):
        emotions.append("恐怖")

    # 怒り・不満
    if re.search(r"(腹立|ムカつ|イライラ|最悪|ひどい)", text, re.IGNORECASE):
        emotions.append("怒り")

    return emotions if emotions else ["中立"]


def classify_theme(text):
    """テーマ分類（ルールベース）"""
    # ノウハウ系
    if re.search(r"(方法|やり方|手順|コツ|ポイント|テクニック|活用|使い方|~する)", text, re.IGNORECASE):
        return "ノウハウ"

    # 実績報告系
    if re.search(r"(達成|突破|記録|フォロワー|収益|売上|成果|結果|~した|~できた)", text, re.IGNORECASE):
        return "実績報告"

    # 体験談系
    if re.search(r"(~してみた|体験|経験|実際|やってみ|試して|使って)", text, re.IGNORECASE):
        return "体験談"

    # ツール紹介系
    if re.search(r"(ツール|AI|ChatGPT|Claude|Grok|アプリ|サービス|おすすめ)", text, re.IGNORECASE):
        return "ツール紹介"

    # 問題提起系
    if re.search(r"(なぜ|どうして|問題|課題|悩み|困っ|\?|？)", text, re.IGNORECASE):
        return "問題提起"

    # 日常系
    if re.search(r"(今日|昨日|明日|朝|昼|夜|ランチ|ご飯|コーヒー)", text, re.IGNORECASE):
        return "日常"

    return "その他"


def detect_url(text):
    """URL有無を検出"""
    return bool(re.search(r"https?://", text))


def detect_special_chars(text):
    """特殊記号の使用を検出"""
    chars = {
        "矢印": bool(re.search(r"[→⇒➡]", text)),
        "括弧": bool(re.search(r"[【】『』「」]", text)),
        "記号": bool(re.search(r"[★☆◆◇■□▼▲]", text)),
    }
    return chars


def count_chars(text):
    """文字数をカウント（空白・改行除く）"""
    return len(re.sub(r"\s", "", text))


def count_line_breaks(text):
    """改行数をカウント"""
    return text.count("\n")


def analyze_posting_time(tweets):
    """投稿時間帯の傾向を分析"""
    hours = []
    for tweet in tweets:
        created_at = tweet.get("createdAt", "")
        if created_at:
            try:
                # ISO 8601形式の日時をパース
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                # 日本時間に変換（+9時間）
                dt_jst = dt + timedelta(hours=9)
                hours.append(dt_jst.hour)
            except Exception:
                continue

    if not hours:
        return {}

    # 時間帯を分類
    morning = sum(1 for h in hours if 6 <= h < 12)  # 6-11時
    afternoon = sum(1 for h in hours if 12 <= h < 18)  # 12-17時
    evening = sum(1 for h in hours if 18 <= h < 24)  # 18-23時
    night = sum(1 for h in hours if 0 <= h < 6)  # 0-5時

    total = len(hours)
    return {
        "朝（6-11時）": f"{morning}件 ({morning/total*100:.1f}%)",
        "昼（12-17時）": f"{afternoon}件 ({afternoon/total*100:.1f}%)",
        "夕方・夜（18-23時）": f"{evening}件 ({evening/total*100:.1f}%)",
        "深夜（0-5時）": f"{night}件 ({night/total*100:.1f}%)",
    }


def calculate_posting_frequency(tweets):
    """投稿頻度を計算（1日あたり何件）"""
    if not tweets:
        return 0

    dates = []
    for tweet in tweets:
        created_at = tweet.get("createdAt", "")
        if created_at:
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                dates.append(dt.date())
            except Exception:
                continue

    if not dates:
        return 0

    # 最古と最新の日付の差を計算
    min_date = min(dates)
    max_date = max(dates)
    days = (max_date - min_date).days + 1

    return len(tweets) / days if days > 0 else 0


def analyze_account(username, tweets):
    """アカウントごとの分析を実行"""
    if not tweets:
        return None

    # エンゲージメント指標を計算
    likes = [tweet.get("likeCount", 0) for tweet in tweets]
    retweets = [tweet.get("retweetCount", 0) for tweet in tweets]
    replies = [tweet.get("replyCount", 0) for tweet in tweets]

    # 各投稿にエンゲージメントスコアと詳細分析を付与
    for tweet in tweets:
        text = tweet.get("text", "")
        tweet["engagement_score"] = (
            tweet.get("likeCount", 0) +
            tweet.get("retweetCount", 0) * 2 +
            tweet.get("replyCount", 0) * 3
        )
        tweet["char_count"] = count_chars(text)
        tweet["line_breaks"] = count_line_breaks(text)
        tweet["opening_pattern"] = classify_opening_pattern(text)
        tweet["has_cta"] = detect_cta(text)
        tweet["emotions"] = analyze_emotion(text)
        tweet["theme"] = classify_theme(text)
        tweet["has_url"] = detect_url(text)
        tweet["special_chars"] = detect_special_chars(text)

    # バズった投稿TOP5とバズらなかった投稿TOP5
    sorted_tweets = sorted(tweets, key=lambda x: x["engagement_score"], reverse=True)
    top5 = sorted_tweets[:5]
    bottom5 = sorted_tweets[-5:] if len(sorted_tweets) >= 5 else sorted_tweets[:1]

    # 各種統計
    opening_patterns = [t["opening_pattern"] for t in tweets]
    pattern_counter = Counter(opening_patterns)

    emotions_all = [e for t in tweets for e in t["emotions"]]
    emotion_counter = Counter(emotions_all)

    themes = [t["theme"] for t in tweets]
    theme_counter = Counter(themes)

    # CTA使用率
    cta_count = sum(1 for t in tweets if t["has_cta"])
    cta_rate = cta_count / len(tweets) * 100 if tweets else 0

    # URL使用率
    url_count = sum(1 for t in tweets if t["has_url"])
    url_rate = url_count / len(tweets) * 100 if tweets else 0

    # 改行数の傾向
    line_breaks = [t["line_breaks"] for t in tweets]

    # 文字数の傾向
    char_counts = [t["char_count"] for t in tweets]

    # 投稿頻度
    posting_freq = calculate_posting_frequency(tweets)

    # 投稿時間帯の傾向
    time_distribution = analyze_posting_time(tweets)

    return {
        "username": username,
        "total_tweets": len(tweets),
        "avg_likes": mean(likes) if likes else 0,
        "median_likes": median(likes) if likes else 0,
        "avg_retweets": mean(retweets) if retweets else 0,
        "median_retweets": median(retweets) if retweets else 0,
        "avg_replies": mean(replies) if replies else 0,
        "median_replies": median(replies) if replies else 0,
        "top5_tweets": top5,
        "bottom5_tweets": bottom5,
        "all_tweets": tweets,  # 全投稿データ（時系列分析用）
        "opening_patterns": dict(pattern_counter),
        "emotions": dict(emotion_counter),
        "themes": dict(theme_counter),
        "cta_rate": cta_rate,
        "url_rate": url_rate,
        "avg_line_breaks": mean(line_breaks) if line_breaks else 0,
        "median_line_breaks": median(line_breaks) if line_breaks else 0,
        "avg_char_count": mean(char_counts) if char_counts else 0,
        "median_char_count": median(char_counts) if char_counts else 0,
        "posting_frequency": posting_freq,
        "time_distribution": time_distribution,
    }


def format_tweet_for_markdown(tweet, rank):
    """投稿をMarkdown形式でフォーマット"""
    text = tweet.get("text", "").replace("\n", " ")[:100]  # 最初の100文字
    likes = tweet.get("likeCount", 0)
    retweets = tweet.get("retweetCount", 0)
    replies = tweet.get("replyCount", 0)
    score = tweet.get("engagement_score", 0)

    username = tweet.get("author", {}).get("userName", "")
    tweet_id = tweet.get("id", "")
    url = f"https://x.com/{username}/status/{tweet_id}" if username and tweet_id else ""

    return f"{rank}. **{text}...** \n   いいね: {likes} | RT: {retweets} | リプ: {replies} | スコア: {score}\n   [{url}]({url})\n"


def compare_top_bottom_posts(analysis):
    """バズった投稿とバズらなかった投稿を比較分析"""
    top5 = analysis["top5_tweets"]
    bottom5 = analysis["bottom5_tweets"]

    def analyze_group(posts):
        """投稿グループの統計を計算"""
        if not posts:
            return {}
        return {
            "avg_chars": mean([p["char_count"] for p in posts]),
            "avg_line_breaks": mean([p["line_breaks"] for p in posts]),
            "opening_patterns": Counter([p["opening_pattern"] for p in posts]),
            "cta_rate": sum(1 for p in posts if p["has_cta"]) / len(posts) * 100,
            "emotions": Counter([e for p in posts for e in p["emotions"]]),
            "themes": Counter([p["theme"] for p in posts]),
            "url_rate": sum(1 for p in posts if p["has_url"]) / len(posts) * 100,
            "arrow_rate": sum(1 for p in posts if p["special_chars"]["矢印"]) / len(posts) * 100,
            "bracket_rate": sum(1 for p in posts if p["special_chars"]["括弧"]) / len(posts) * 100,
            "symbol_rate": sum(1 for p in posts if p["special_chars"]["記号"]) / len(posts) * 100,
        }

    top_stats = analyze_group(top5)
    bottom_stats = analyze_group(bottom5)

    return {
        "top": top_stats,
        "bottom": bottom_stats,
    }


def analyze_cross_account_patterns(all_analyses):
    """アカウント間の横断分析"""
    if not all_analyses:
        return {}

    # エンゲージメント率ランキング
    rankings = {
        "likes": sorted(all_analyses, key=lambda x: x["avg_likes"], reverse=True),
        "retweets": sorted(all_analyses, key=lambda x: x["avg_retweets"], reverse=True),
        "replies": sorted(all_analyses, key=lambda x: x["avg_replies"], reverse=True),
    }

    # 投稿頻度とエンゲージメントの関係
    freq_engagement = [
        {
            "username": a["username"],
            "posting_freq": a["posting_frequency"],
            "avg_likes": a["avg_likes"],
        }
        for a in all_analyses
    ]

    # 各アカウントの得意パターン
    strong_patterns = []
    for analysis in all_analyses:
        top5 = analysis["top5_tweets"]
        if top5:
            top_patterns = Counter([t["opening_pattern"] for t in top5])
            top_themes = Counter([t["theme"] for t in top5])
            top_emotions = Counter([e for t in top5 for e in t["emotions"]])

            strong_patterns.append({
                "username": analysis["username"],
                "best_opening": top_patterns.most_common(1)[0][0] if top_patterns else "不明",
                "best_theme": top_themes.most_common(1)[0][0] if top_themes else "不明",
                "best_emotion": top_emotions.most_common(1)[0][0] if top_emotions else "不明",
            })

    return {
        "rankings": rankings,
        "freq_engagement": freq_engagement,
        "strong_patterns": strong_patterns,
    }


def generate_recommendations(all_analyses):
    """拓巳への具体的提案を生成"""
    if not all_analyses:
        return {"should_do": [], "should_avoid": []}

    # 全アカウントのTOP5投稿を集計
    all_top_tweets = []
    for analysis in all_analyses:
        all_top_tweets.extend(analysis["top5_tweets"])

    # 全アカウントのBOTTOM5投稿を集計
    all_bottom_tweets = []
    for analysis in all_analyses:
        all_bottom_tweets.extend(analysis["bottom5_tweets"])

    # バズった投稿の共通点を分析
    if all_top_tweets:
        top_patterns = Counter([t["opening_pattern"] for t in all_top_tweets])
        top_themes = Counter([t["theme"] for t in all_top_tweets])
        top_emotions = Counter([e for t in all_top_tweets for e in t["emotions"]])
        top_cta_rate = sum(1 for t in all_top_tweets if t["has_cta"]) / len(all_top_tweets) * 100
        top_url_rate = sum(1 for t in all_top_tweets if t["has_url"]) / len(all_top_tweets) * 100
        top_avg_chars = mean([t["char_count"] for t in all_top_tweets])
        top_avg_line_breaks = mean([t["line_breaks"] for t in all_top_tweets])

    # バズらなかった投稿の共通点を分析
    if all_bottom_tweets:
        bottom_patterns = Counter([t["opening_pattern"] for t in all_bottom_tweets])
        bottom_themes = Counter([t["theme"] for t in all_bottom_tweets])
        bottom_cta_rate = sum(1 for t in all_bottom_tweets if t["has_cta"]) / len(all_bottom_tweets) * 100
        bottom_avg_chars = mean([t["char_count"] for t in all_bottom_tweets])

    # 提案を生成
    should_do = []
    should_avoid = []

    # 真似すべきパターン（優先順位付き）
    if all_top_tweets:
        # 1. 冒頭パターン
        best_pattern = top_patterns.most_common(1)[0]
        should_do.append({
            "priority": 1,
            "action": f"冒頭は「{best_pattern[0]}」パターンを使う",
            "reason": f"バズった投稿の{best_pattern[1]/len(all_top_tweets)*100:.1f}%がこのパターンを使用",
        })

        # 2. テーマ
        best_theme = top_themes.most_common(1)[0]
        should_do.append({
            "priority": 2,
            "action": f"「{best_theme[0]}」系の投稿を増やす",
            "reason": f"バズった投稿の{best_theme[1]/len(all_top_tweets)*100:.1f}%がこのテーマ",
        })

        # 3. 感情
        best_emotion = top_emotions.most_common(1)[0]
        should_do.append({
            "priority": 3,
            "action": f"「{best_emotion[0]}」の感情を呼び起こす表現を使う",
            "reason": f"バズった投稿で最も多い感情（{best_emotion[1]}回出現）",
        })

        # 4. 文字数
        should_do.append({
            "priority": 4,
            "action": f"文字数は{top_avg_chars:.0f}文字前後を目安にする",
            "reason": f"バズった投稿の平均文字数",
        })

        # 5. CTAまたはURL
        if top_cta_rate > 40:
            should_do.append({
                "priority": 5,
                "action": "CTAを積極的に入れる",
                "reason": f"バズった投稿の{top_cta_rate:.1f}%がCTAを使用",
            })
        elif top_url_rate > 30:
            should_do.append({
                "priority": 5,
                "action": "URLを含める（参考リンクや自分のコンテンツ）",
                "reason": f"バズった投稿の{top_url_rate:.1f}%がURLを含む",
            })
        else:
            should_do.append({
                "priority": 5,
                "action": f"改行を{top_avg_line_breaks:.0f}回程度使って読みやすくする",
                "reason": f"バズった投稿の平均改行数",
            })

    # 避けるべきパターン
    if all_bottom_tweets:
        # 1. バズらなかったパターン
        worst_pattern = bottom_patterns.most_common(1)[0]
        should_avoid.append({
            "action": f"「{worst_pattern[0]}」パターンは避ける",
            "reason": f"バズらなかった投稿の{worst_pattern[1]/len(all_bottom_tweets)*100:.1f}%がこのパターン",
        })

        # 2. バズらなかったテーマ
        worst_theme = bottom_themes.most_common(1)[0]
        should_avoid.append({
            "action": f"「{worst_theme[0]}」系の投稿は控えめにする",
            "reason": f"バズらなかった投稿の{worst_theme[1]/len(all_bottom_tweets)*100:.1f}%がこのテーマ",
        })

        # 3. 文字数
        if bottom_avg_chars < top_avg_chars * 0.7:
            should_avoid.append({
                "action": f"{bottom_avg_chars:.0f}文字以下の短すぎる投稿は避ける",
                "reason": "バズらなかった投稿の平均文字数が少なすぎる",
            })
        else:
            should_avoid.append({
                "action": f"CTAなしの投稿を連投しない",
                "reason": f"バズらなかった投稿のCTA使用率は{bottom_cta_rate:.1f}%のみ",
            })

    return {
        "should_do": should_do[:5],
        "should_avoid": should_avoid[:3],
    }


def generate_markdown_report(all_analyses, output_path):
    """Markdown形式の詳細レポートを生成"""
    today = datetime.now().strftime("%Y年%m月%d日")

    md_content = f"""# Xアカウント分析レポート（詳細版）

**生成日時**: {today}
**分析アカウント数**: {len(all_analyses)}個

---

"""

    # アカウント間比較分析
    if len(all_analyses) > 1:
        md_content += "## 📊 全アカウント横断比較\n\n"
        cross_analysis = analyze_cross_account_patterns(all_analyses)

        # エンゲージメント率ランキング
        md_content += "### エンゲージメント率ランキング\n\n"
        md_content += "#### 平均いいね数ランキング\n\n"
        for i, analysis in enumerate(cross_analysis["rankings"]["likes"], 1):
            md_content += f"{i}. **@{analysis['username']}**: {analysis['avg_likes']:.1f}件\n"
        md_content += "\n"

        md_content += "#### 平均RT数ランキング\n\n"
        for i, analysis in enumerate(cross_analysis["rankings"]["retweets"], 1):
            md_content += f"{i}. **@{analysis['username']}**: {analysis['avg_retweets']:.1f}件\n"
        md_content += "\n"

        md_content += "#### 平均リプライ数ランキング\n\n"
        for i, analysis in enumerate(cross_analysis["rankings"]["replies"], 1):
            md_content += f"{i}. **@{analysis['username']}**: {analysis['avg_replies']:.1f}件\n"
        md_content += "\n"

        # 投稿頻度とエンゲージメントの関係
        md_content += "### 投稿頻度とエンゲージメントの関係\n\n"
        md_content += "| アカウント | 投稿頻度（件/日） | 平均いいね数 |\n"
        md_content += "|-----------|----------------|-------------|\n"
        for item in cross_analysis["freq_engagement"]:
            md_content += f"| @{item['username']} | {item['posting_freq']:.2f} | {item['avg_likes']:.1f} |\n"
        md_content += "\n"

        # 各アカウントの得意パターン
        md_content += "### 各アカウントの得意パターン一覧\n\n"
        md_content += "| アカウント | 効果的な冒頭 | 効果的なテーマ | 効果的な感情 |\n"
        md_content += "|-----------|------------|--------------|------------|\n"
        for item in cross_analysis["strong_patterns"]:
            md_content += f"| @{item['username']} | {item['best_opening']} | {item['best_theme']} | {item['best_emotion']} |\n"
        md_content += "\n---\n\n"

    # 拓巳への具体的提案
    recommendations = generate_recommendations(all_analyses)
    md_content += "## 💡 拓巳への具体的提案\n\n"
    md_content += "### ✅ 真似すべきパターン（優先順位順）\n\n"
    for rec in recommendations["should_do"]:
        md_content += f"**{rec['priority']}. {rec['action']}**\n"
        md_content += f"   - 根拠: {rec['reason']}\n\n"

    md_content += "### ❌ 避けるべきパターン\n\n"
    for rec in recommendations["should_avoid"]:
        md_content += f"- **{rec['action']}**\n"
        md_content += f"  - 理由: {rec['reason']}\n\n"

    md_content += "---\n\n"

    # 各アカウントの詳細分析
    for analysis in all_analyses:
        username = analysis["username"]

        md_content += f"## 👤 @{username}\n\n"

        # 基本統計
        md_content += "### 📈 基本統計\n\n"
        md_content += f"- **投稿数**: {analysis['total_tweets']}件\n"
        md_content += f"- **平均いいね数**: {analysis['avg_likes']:.1f} （中央値: {analysis['median_likes']:.0f}）\n"
        md_content += f"- **平均RT数**: {analysis['avg_retweets']:.1f} （中央値: {analysis['median_retweets']:.0f}）\n"
        md_content += f"- **平均リプライ数**: {analysis['avg_replies']:.1f} （中央値: {analysis['median_replies']:.0f}）\n"
        md_content += f"- **平均文字数**: {analysis['avg_char_count']:.0f}文字 （中央値: {analysis['median_char_count']:.0f}文字）\n"
        md_content += f"- **投稿頻度**: {analysis['posting_frequency']:.2f}件/日\n\n"

        # 投稿時間帯
        if analysis["time_distribution"]:
            md_content += "### ⏰ 投稿時間帯の傾向\n\n"
            for time_slot, value in analysis["time_distribution"].items():
                md_content += f"- {time_slot}: {value}\n"
            md_content += "\n"

        # バズった vs バズらなかった投稿の比較分析
        comparison = compare_top_bottom_posts(analysis)
        md_content += "### 🔥 vs 📉 バズった投稿 vs バズらなかった投稿の比較\n\n"
        md_content += "| 項目 | TOP5 | BOTTOM5 | 差分 |\n"
        md_content += "|------|------|---------|------|\n"

        top = comparison["top"]
        bottom = comparison["bottom"]

        md_content += f"| 平均文字数 | {top['avg_chars']:.0f}文字 | {bottom['avg_chars']:.0f}文字 | {top['avg_chars'] - bottom['avg_chars']:+.0f} |\n"
        md_content += f"| 平均改行数 | {top['avg_line_breaks']:.1f}回 | {bottom['avg_line_breaks']:.1f}回 | {top['avg_line_breaks'] - bottom['avg_line_breaks']:+.1f} |\n"
        md_content += f"| CTA使用率 | {top['cta_rate']:.0f}% | {bottom['cta_rate']:.0f}% | {top['cta_rate'] - bottom['cta_rate']:+.0f}% |\n"
        md_content += f"| URL使用率 | {top['url_rate']:.0f}% | {bottom['url_rate']:.0f}% | {top['url_rate'] - bottom['url_rate']:+.0f}% |\n"
        md_content += f"| 矢印使用率 | {top['arrow_rate']:.0f}% | {bottom['arrow_rate']:.0f}% | {top['arrow_rate'] - bottom['arrow_rate']:+.0f}% |\n"
        md_content += f"| 括弧使用率 | {top['bracket_rate']:.0f}% | {bottom['bracket_rate']:.0f}% | {top['bracket_rate'] - bottom['bracket_rate']:+.0f}% |\n"
        md_content += f"| 記号使用率 | {top['symbol_rate']:.0f}% | {bottom['symbol_rate']:.0f}% | {top['symbol_rate'] - bottom['symbol_rate']:+.0f}% |\n"
        md_content += "\n"

        # 冒頭パターンの比較
        md_content += "**冒頭パターン比較:**\n\n"
        md_content += "- TOP5: " + ", ".join([f"{k}({v}件)" for k, v in top["opening_patterns"].most_common(3)]) + "\n"
        md_content += "- BOTTOM5: " + ", ".join([f"{k}({v}件)" for k, v in bottom["opening_patterns"].most_common(3)]) + "\n\n"

        # 感情の比較
        md_content += "**感情の比較:**\n\n"
        md_content += "- TOP5: " + ", ".join([f"{k}({v}回)" for k, v in top["emotions"].most_common(3)]) + "\n"
        md_content += "- BOTTOM5: " + ", ".join([f"{k}({v}回)" for k, v in bottom["emotions"].most_common(3)]) + "\n\n"

        # テーマの比較
        md_content += "**テーマの比較:**\n\n"
        md_content += "- TOP5: " + ", ".join([f"{k}({v}件)" for k, v in top["themes"].most_common(3)]) + "\n"
        md_content += "- BOTTOM5: " + ", ".join([f"{k}({v}件)" for k, v in bottom["themes"].most_common(3)]) + "\n\n"

        # バズった投稿TOP5
        md_content += "### 🔥 最もバズった投稿 TOP5\n\n"
        for i, tweet in enumerate(analysis["top5_tweets"], 1):
            md_content += format_tweet_for_markdown(tweet, i)
        md_content += "\n"

        # バズらなかった投稿BOTTOM5
        md_content += "### 📉 最もバズらなかった投稿 BOTTOM5\n\n"
        for i, tweet in enumerate(analysis["bottom5_tweets"], 1):
            md_content += format_tweet_for_markdown(tweet, i)
        md_content += "\n"

        # 時系列分析（簡易版）
        all_tweets = analysis["all_tweets"]
        if all_tweets:
            # 日時でソート
            sorted_by_time = sorted(all_tweets, key=lambda x: x.get("createdAt", ""))

            md_content += "### 📈 時系列分析（いいね数の推移）\n\n"
            md_content += "直近10件の投稿のいいね数推移:\n\n"
            md_content += "| 順位 | 投稿日時 | いいね数 | 本文（抜粋） |\n"
            md_content += "|------|---------|---------|-------------|\n"

            for i, tweet in enumerate(sorted_by_time[-10:], 1):
                created = tweet.get("createdAt", "")[:10] if tweet.get("createdAt") else "不明"
                likes = tweet.get("likeCount", 0)
                text = tweet.get("text", "")[:30].replace("\n", " ")
                md_content += f"| {i} | {created} | {likes} | {text}... |\n"
            md_content += "\n"

            # トレンド分析
            if len(sorted_by_time) >= 10:
                recent_5 = sorted_by_time[-5:]
                older_5 = sorted_by_time[-10:-5]
                recent_avg = mean([t.get("likeCount", 0) for t in recent_5])
                older_avg = mean([t.get("likeCount", 0) for t in older_5])
                trend = "上昇傾向" if recent_avg > older_avg else "下降傾向" if recent_avg < older_avg else "横ばい"

                md_content += f"**トレンド**: {trend} （直近5件平均: {recent_avg:.1f}、その前5件平均: {older_avg:.1f}）\n\n"

        # 冒頭パターン分析
        md_content += "### 🎯 冒頭パターン分析\n\n"
        total_patterns = sum(analysis["opening_patterns"].values())
        for pattern, count in sorted(analysis["opening_patterns"].items(), key=lambda x: x[1], reverse=True):
            percentage = count / total_patterns * 100 if total_patterns > 0 else 0
            md_content += f"- **{pattern}**: {count}件 ({percentage:.1f}%)\n"
        md_content += "\n"

        # 感情分析
        md_content += "### 😊 感情分析\n\n"
        total_emotions = sum(analysis["emotions"].values())
        for emotion, count in sorted(analysis["emotions"].items(), key=lambda x: x[1], reverse=True):
            percentage = count / total_emotions * 100 if total_emotions > 0 else 0
            md_content += f"- **{emotion}**: {count}回 ({percentage:.1f}%)\n"
        md_content += "\n"

        # テーマ分類
        md_content += "### 📚 テーマ分類\n\n"
        total_themes = sum(analysis["themes"].values())
        for theme, count in sorted(analysis["themes"].items(), key=lambda x: x[1], reverse=True):
            percentage = count / total_themes * 100 if total_themes > 0 else 0
            md_content += f"- **{theme}**: {count}件 ({percentage:.1f}%)\n"
        md_content += "\n"

        # その他の統計
        md_content += "### 📊 その他の統計\n\n"
        md_content += f"- **CTA使用率**: {analysis['cta_rate']:.1f}%\n"
        md_content += f"- **URL使用率**: {analysis['url_rate']:.1f}%\n"
        md_content += f"- **平均改行数**: {analysis['avg_line_breaks']:.1f} （中央値: {analysis['median_line_breaks']:.0f}）\n\n"

        md_content += "---\n\n"

    # ファイルに保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"詳細レポート保存完了: {output_path}")


def save_to_excel(all_tweets_by_account, output_path):
    """全投稿データをExcelに保存（アカウントごとにシート分け）"""
    wb = Workbook()

    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for username, tweets in all_tweets_by_account.items():
        # シート名（最大31文字制限）
        sheet_name = f"@{username}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # ヘッダー
        headers = ["本文", "いいね数", "RT数", "リプライ数", "投稿日時", "投稿URL", "冒頭パターン", "CTA有無", "改行数"]
        ws.append(headers)

        # ヘッダー行のスタイル設定
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # データ行を追加
        for tweet in tweets:
            text = tweet.get("text", "")
            tweet_id = tweet.get("id", "")
            post_url = f"https://x.com/{username}/status/{tweet_id}" if tweet_id else ""

            ws.append([
                text,
                tweet.get("likeCount", 0),
                tweet.get("retweetCount", 0),
                tweet.get("replyCount", 0),
                tweet.get("createdAt", ""),
                post_url,
                classify_opening_pattern(text),
                "あり" if detect_cta(text) else "なし",
                count_line_breaks(text),
            ])

        # 本文列（A列）の折り返し設定
        for row in range(2, len(tweets) + 2):
            cell = ws.cell(row=row, column=1)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 列幅の調整
        column_widths = {
            1: 60,  # 本文
            2: 12,  # いいね数
            3: 12,  # RT数
            4: 12,  # リプライ数
            5: 25,  # 投稿日時
            6: 50,  # 投稿URL
            7: 15,  # 冒頭パターン
            8: 12,  # CTA有無
            9: 10,  # 改行数
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # オートフィルター設定
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(tweets) + 1}"

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel保存完了: {output_path}")


def main(test_mode=False, test_file=None):
    """メイン処理"""
    print("=" * 60)
    print("Xアカウント分析スクリプト")
    if test_mode:
        print("【テストモード】")
    print("=" * 60)
    print()

    # 各アカウントの投稿を取得
    all_tweets_by_account = {}
    all_analyses = []

    if test_mode:
        # テストモード: 既存Excelファイルから読み込み
        if not test_file:
            test_file = "output/buzz_posts_20260215.xlsx"

        all_tweets_by_account = load_test_data_from_excel(test_file)
        print()

        if not all_tweets_by_account:
            print("エラー: テストデータが読み込めませんでした。")
            sys.exit(1)

        # アカウントごとに分析
        print("分析を実行中...")
        for username, tweets in all_tweets_by_account.items():
            print(f"  @{username} を分析中...")
            analysis = analyze_account(username, tweets)
            if analysis:
                all_analyses.append(analysis)
        print()

    else:
        # APIモード: 通常のAPI取得
        api_key = os.environ.get("TWITTER_API_KEY")
        if not api_key:
            print("エラー: 環境変数 TWITTER_API_KEY が設定されていません。")
            print(".envファイルの読み込みを確認してください。")
            sys.exit(1)

        print(f"対象アカウント数: {len(TARGET_ACCOUNTS)}個")
        print()

        for i, username in enumerate(TARGET_ACCOUNTS):
            print(f"[{i+1}/{len(TARGET_ACCOUNTS)}] @{username} を処理中...")

            tweets = fetch_user_tweets(username, api_key, count=100)

            if tweets:
                all_tweets_by_account[username] = tweets

                # アカウント分析を実行
                analysis = analyze_account(username, tweets)
                if analysis:
                    all_analyses.append(analysis)

            # レート制限対策：最後のアカウント以外は待機
            if i < len(TARGET_ACCOUNTS) - 1:
                print("    レート制限対策で10秒待機中...")
                time.sleep(10)

            print()

    if not all_analyses:
        print("分析可能なデータが取得できませんでした。")
        sys.exit(0)

    print(f"分析完了: {len(all_analyses)}アカウント")
    print()

    # 出力ファイルパス
    today = datetime.now().strftime("%Y%m%d")
    output_dir = "output"
    md_path = os.path.join(output_dir, f"account_analysis_詳細_{today}.md")
    excel_path = os.path.join(output_dir, f"account_posts_{today}.xlsx")

    # Markdownレポート生成
    print("Markdownレポートを生成中...")
    generate_markdown_report(all_analyses, md_path)

    # Excel保存（テストモードでは既存データを保存）
    if not test_mode:
        print("Excelファイルを生成中...")
        save_to_excel(all_tweets_by_account, excel_path)
    else:
        print("（テストモードのため、Excel生成はスキップ）")

    print()
    print("=" * 60)
    print("すべての処理が完了しました！")
    print("=" * 60)


if __name__ == "__main__":
    # コマンドライン引数の解析
    parser = argparse.ArgumentParser(description="Xアカウント分析スクリプト")
    parser.add_argument(
        "--test",
        action="store_true",
        help="テストモード（既存Excelファイルから読み込み、API呼び出しなし）",
    )
    parser.add_argument(
        "--test-file",
        type=str,
        default="output/buzz_posts_20260215.xlsx",
        help="テストモードで使用するExcelファイルのパス",
    )

    args = parser.parse_args()

    main(test_mode=args.test, test_file=args.test_file)
